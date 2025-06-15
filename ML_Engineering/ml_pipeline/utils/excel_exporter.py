"""
Excel Exporter module for generating comprehensive Excel reports.

This module creates detailed Excel files with multiple sheets containing
ML pipeline results, feature analysis, and evaluation metrics.
"""

import logging
import pandas as pd
import os
from typing import List, Dict, Any, Optional
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    openpyxl = None

# Configure logging
logger = logging.getLogger(__name__)

class ExcelExporter:
    """
    Excel exporter for ML pipeline results.
    """
    
    def __init__(self):
        """Initialize the Excel exporter."""
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl not available. Excel export will be limited.")
    
    def export_ml_results(self, results: Dict[str, Any], 
                         output_path: str,
                         include_raw_data: bool = True) -> bool:
        """
        Export ML pipeline results to Excel file.
        
        Args:
            results: ML pipeline results dictionary
            output_path: Path to save Excel file
            include_raw_data: Whether to include raw segment data
            
        Returns:
            True if successful, False otherwise
        """
        try:
            if not OPENPYXL_AVAILABLE:
                logger.warning("openpyxl not available - using basic CSV export")
                return self._export_basic_csv(results, output_path)
            
            logger.info(f"Exporting ML results to Excel: {output_path}")
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create sheets
            self._create_summary_sheet(wb, results)
            self._create_segments_sheet(wb, results, include_raw_data)
            self._create_features_sheet(wb, results)
            self._create_evaluation_sheet(wb, results)
            
            if results.get("topic_modeling_results"):
                self._create_topics_sheet(wb, results)
            
            # Save workbook
            wb.save(output_path)
            logger.info(f"Excel file saved successfully: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            return False
    
    def _create_summary_sheet(self, wb: Any, results: Dict[str, Any]):
        """Create summary sheet with overview information."""
        ws = wb.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = "ML Pipeline Results Summary"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        
        # Merge title cells
        ws.merge_cells('A1:D1')
        
        # Basic information
        row = 3
        info_data = [
            ("Processing Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Total Segments Processed", results.get("input_segments", 0)),
            ("Features Enhanced", "Yes" if results.get("feature_engineering_results", {}).get("features_added") else "No"),
            ("Topic Modeling", "Available" if results.get("topic_modeling_results") else "Not Available"),
            ("Semantic Search", "Available" if results.get("semantic_search_results", {}).get("index_built") else "Not Available"),
            ("Evaluation Completed", "Yes" if results.get("evaluation_results") else "No")
        ]
        
        for label, value in info_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        # Evaluation summary
        if results.get("evaluation_results", {}).get("summary"):
            row += 1
            ws[f'A{row}'] = "Evaluation Summary"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            eval_summary = results["evaluation_results"]["summary"]
            for key, value in eval_summary.items():
                ws[f'A{row}'] = key.replace("_", " ").title()
                ws[f'B{row}'] = value if isinstance(value, (str, int, float)) else str(value)
                row += 1
        
        # Auto-adjust column widths
        for col_idx in range(1, 5):  # Adjust first 4 columns
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value and not isinstance(cell, openpyxl.cell.MergedCell):
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_segments_sheet(self, wb: Any, results: Dict[str, Any],
                              include_raw_data: bool):
        """Create segments sheet with processed segment data."""
        ws = wb.create_sheet("Segments")
        
        segments = results.get("processed_segments", [])
        if not segments:
            ws['A1'] = "No segments data available"
            return
        
        # Prepare data for DataFrame
        segment_data = []
        
        for segment in segments:
            row_data = {
                "Segment ID": segment.get("segment_id", ""),
                "Source Document": segment.get("source_doc_id", ""),
                "Text Length": len(segment.get("text", "")),
                "Word Count": segment.get("features", {}).get("word_count", 0),
                "Sentence Count": segment.get("features", {}).get("sentence_count", 0),
                "Temporal Context": segment.get("features", {}).get("temporal_context", "unknown"),
                "Sustainability Score": segment.get("features", {}).get("total_sustainability_score", 0),
                "Discourse Markers": len(segment.get("features", {}).get("discourse_markers", [])),
                "Noun Phrases": len(segment.get("features", {}).get("noun_phrases", [])),
                "Lexical Diversity": segment.get("features", {}).get("lexical_diversity", 0.0)
            }
            
            # Add topic modeling results if available
            if segment.get("topic_modeling"):
                row_data["Topic ID"] = segment["topic_modeling"].get("topic_id", -1)
                row_data["Topic Probability"] = segment["topic_modeling"].get("probability", 0.0)
            
            # Add raw text if requested
            if include_raw_data:
                text = segment.get("text", "")
                row_data["Text"] = text[:500] + "..." if len(text) > 500 else text
            
            segment_data.append(row_data)
        
        # Create DataFrame and write to sheet
        df = pd.DataFrame(segment_data)
        
        # Write headers
        for col_idx, column in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Write data
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for col_idx in range(1, len(df.columns) + 1):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
            adjusted_width = min(max_length + 2, 80)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_features_sheet(self, wb: Any, results: Dict[str, Any]):
        """Create features analysis sheet."""
        ws = wb.create_sheet("Features Analysis")
        
        segments = results.get("processed_segments", [])
        if not segments:
            ws['A1'] = "No features data available"
            return
        
        # Collect all feature statistics
        feature_stats = {}
        
        for segment in segments:
            features = segment.get("features", {})
            for feature_name, feature_value in features.items():
                if isinstance(feature_value, (int, float)):
                    if feature_name not in feature_stats:
                        feature_stats[feature_name] = []
                    feature_stats[feature_name].append(feature_value)
        
        # Calculate statistics
        stats_data = []
        for feature_name, values in feature_stats.items():
            if values:
                stats_data.append({
                    "Feature": feature_name.replace("_", " ").title(),
                    "Count": len(values),
                    "Mean": round(sum(values) / len(values), 3),
                    "Min": min(values),
                    "Max": max(values),
                    "Std Dev": round(pd.Series(values).std(), 3) if len(values) > 1 else 0
                })
        
        # Create DataFrame and write to sheet
        if stats_data:
            df = pd.DataFrame(stats_data)
            
            # Write headers
            for col_idx, column in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=column)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            
            # Write data
            for row_idx, row in enumerate(df.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        if stats_data:
            for col_idx in range(1, len(df.columns) + 1):
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(col_idx)
                for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_evaluation_sheet(self, wb: Any, results: Dict[str, Any]):
        """Create evaluation metrics sheet."""
        ws = wb.create_sheet("Evaluation")
        
        eval_results = results.get("evaluation_results", {})
        if not eval_results:
            ws['A1'] = "No evaluation data available"
            return
        
        row = 1
        
        # Summary metrics
        if eval_results.get("summary"):
            ws[f'A{row}'] = "Summary Metrics"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 2
            
            for key, value in eval_results["summary"].items():
                ws[f'A{row}'] = key.replace("_", " ").title()
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = value if isinstance(value, (str, int, float)) else str(value)
                row += 1
            
            row += 1
        
        # Detailed metrics
        if eval_results.get("detailed_metrics"):
            ws[f'A{row}'] = "Detailed Metrics"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 2
            
            detailed = eval_results["detailed_metrics"]
            
            # Feature quality metrics
            if detailed.get("feature_quality"):
                ws[f'A{row}'] = "Feature Quality"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                
                fq = detailed["feature_quality"]
                quality_data = [
                    ("Total Segments", fq.get("total_segments", 0)),
                    ("Data Quality Score", fq.get("data_quality", {}).get("feature_completeness_ratio", 0))
                ]
                
                for label, value in quality_data:
                    ws[f'B{row}'] = label
                    ws[f'C{row}'] = value
                    row += 1
                
                row += 1
    
    def _create_topics_sheet(self, wb: Any, results: Dict[str, Any]):
        """Create topics analysis sheet."""
        ws = wb.create_sheet("Topics")
        
        topic_results = results.get("topic_modeling_results", {})
        topics = topic_results.get("topics", {})
        
        if not topics:
            ws['A1'] = "No topic modeling data available"
            return
        
        # Prepare topics data
        topics_data = []
        
        for topic_id, topic_info in topics.items():
            keywords = topic_info.get("keywords", [])
            keyword_terms = [kw.get("term", "") for kw in keywords[:5]]  # Top 5 keywords
            
            topics_data.append({
                "Topic ID": topic_id,
                "Document Count": topic_info.get("count", 0),
                "Topic Name": topic_info.get("name", f"Topic {topic_id}"),
                "Top Keywords": ", ".join(keyword_terms),
                "Keyword Count": len(keywords)
            })
        
        # Create DataFrame and write to sheet
        if topics_data:
            df = pd.DataFrame(topics_data)
            
            # Write headers
            for col_idx, column in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=column)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            
            # Write data
            for row_idx, row in enumerate(df.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        if topics_data:
            for col_idx in range(1, len(df.columns) + 1):
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(col_idx)
                for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                adjusted_width = min(max_length + 2, 60)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def _export_basic_csv(self, results: Dict[str, Any], output_path: str) -> bool:
        """Fallback CSV export when openpyxl is not available."""
        try:
            segments = results.get("processed_segments", [])
            if not segments:
                logger.warning("No segments data to export")
                return False
            
            # Prepare data for CSV
            csv_data = []
            for segment in segments:
                features = segment.get("features", {})
                row_data = {
                    "segment_id": segment.get("segment_id", ""),
                    "source_doc_id": segment.get("source_doc_id", ""),
                    "text": segment.get("text", ""),
                    "word_count": features.get("word_count", 0),
                    "sentence_count": features.get("sentence_count", 0),
                    "temporal_context": features.get("temporal_context", "unknown"),
                    "sustainability_score": features.get("total_sustainability_score", 0)
                }
                csv_data.append(row_data)
            
            # Save as CSV
            df = pd.DataFrame(csv_data)
            csv_path = output_path.replace('.xlsx', '.csv')
            df.to_csv(csv_path, index=False, encoding='utf-8')
            
            logger.info(f"Basic CSV export completed: {csv_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error in basic CSV export: {e}")
            return False
